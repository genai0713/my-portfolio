from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from random import Random

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties


BASE_DIR = Path(__file__).resolve().parents[1]
OUTPUT_DIR = BASE_DIR / "output" / "spreadsheet"
OUTPUT_PATH = OUTPUT_DIR / "sales_dashboard.xlsx"


def sample_sales_data() -> list[dict[str, object]]:
    rng = Random(42)
    now = datetime.now().replace(hour=10, minute=0, second=0, microsecond=0)
    senders = ["Aisha", "Rahul", "Fatima", "Noah", "Priya", "Omar"]
    products = [
        ("iPhone 15", 49999),
        ("Galaxy S24", 45999),
        ("AirPods Pro", 18999),
        ("Apple Watch", 22999),
        ("iPad Air", 32999),
        ("MacBook Air", 84999),
    ]
    customers = [
        "John Doe",
        "Meera Shah",
        "Ali Khan",
        "Sara James",
        "Rohan Patel",
        "Nora Salem",
        "Kabir Singh",
        "Emma Clark",
        "Yusuf Ahmed",
        "Anya Verma",
        "Leah Thomas",
        "Arjun Mehta",
    ]

    rows: list[dict[str, object]] = []
    for index in range(1, 85):
        product_name, base_amount = rng.choice(products)
        captured_at = now - timedelta(days=(84 - index) * 2 + rng.randint(0, 1))
        discount = rng.choice([0, 0, 0, 1200, 1800, 2500])
        amount = base_amount - discount
        parse_status = "parsed" if rng.random() > 0.08 else "failed"
        customer = rng.choice(customers) if parse_status == "parsed" else ""
        phone = f"9{rng.randint(100000000, 999999999)}"

        rows.append(
            {
                "captured_at": captured_at,
                "group_name": "Dubai Retail Sales",
                "sender": rng.choice(senders),
                "message_id": f"wamid-{index:04d}",
                "raw_text": (
                    f"Sale: customer={customer or 'unknown'}, product={product_name}, "
                    f"amount={amount}, phone={phone}"
                ),
                "customer": customer,
                "product": product_name if parse_status == "parsed" else "",
                "amount": amount if parse_status == "parsed" else 0,
                "phone": phone if parse_status == "parsed" else "",
                "parse_status": parse_status,
            }
        )
    return rows


def apply_sheet_title(sheet, title: str) -> None:
    sheet.merge_cells("A1:H1")
    cell = sheet["A1"]
    cell.value = title
    cell.font = Font(size=16, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="0B5D6B")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 24


def build_raw_data_sheet(workbook: Workbook, rows: list[dict[str, object]]) -> None:
    sheet = workbook.active
    sheet.title = "RawData"

    headers = [
        "captured_at",
        "group_name",
        "sender",
        "message_id",
        "raw_text",
        "customer",
        "product",
        "amount",
        "phone",
        "parse_status",
        "order_month",
        "customer_key",
        "first_customer_flag",
    ]
    sheet.append(headers)

    for row_index, row in enumerate(rows, start=2):
        sheet.cell(row=row_index, column=1, value=row["captured_at"])
        sheet.cell(row=row_index, column=2, value=row["group_name"])
        sheet.cell(row=row_index, column=3, value=row["sender"])
        sheet.cell(row=row_index, column=4, value=row["message_id"])
        sheet.cell(row=row_index, column=5, value=row["raw_text"])
        sheet.cell(row=row_index, column=6, value=row["customer"])
        sheet.cell(row=row_index, column=7, value=row["product"])
        sheet.cell(row=row_index, column=8, value=row["amount"])
        sheet.cell(row=row_index, column=9, value=row["phone"])
        sheet.cell(row=row_index, column=10, value=row["parse_status"])
        sheet.cell(
            row=row_index,
            column=11,
            value=f'=IF(A{row_index}="","",DATE(YEAR(A{row_index}),MONTH(A{row_index}),1))',
        )
        sheet.cell(
            row=row_index,
            column=12,
            value=f'=IF(F{row_index}<>"",LOWER(TRIM(F{row_index})),D{row_index})',
        )
        sheet.cell(
            row=row_index,
            column=13,
            value=(
                f'=IF(L{row_index}="","",IF(COUNTIF($L$2:L{row_index},'
                f'L{row_index})=1,1,0))'
            ),
        )

    header_fill = PatternFill("solid", fgColor="D9E7EB")
    input_font = Font(color="0000FF", bold=True)
    derived_font = Font(color="000000")

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index in range(2, len(rows) + 2):
        for column in range(1, 11):
            sheet.cell(row=row_index, column=column).font = input_font
        for column in range(11, 14):
            sheet.cell(row=row_index, column=column).font = derived_font

    currency_format = '"$"#,##0;[Red]("$"#,##0)'
    date_format = "dd-mmm-yyyy hh:mm"
    month_format = "mmm-yy"
    for row_index in range(2, len(rows) + 2):
        sheet.cell(row=row_index, column=1).number_format = date_format
        sheet.cell(row=row_index, column=8).number_format = currency_format
        sheet.cell(row=row_index, column=11).number_format = month_format

    widths = {
        "A": 20,
        "B": 20,
        "C": 14,
        "D": 16,
        "E": 52,
        "F": 18,
        "G": 18,
        "H": 14,
        "I": 16,
        "J": 14,
        "K": 14,
        "L": 18,
        "M": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    table = Table(displayName="SalesData", ref=f"A1:M{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.sheet_view.showGridLines = False

    source_comment = Comment(
        "Example dashboard data generated from the local project schema. Replace with live Google Sheets exports when available.",
        "Codex",
    )
    sheet["A1"].comment = source_comment


def build_summary_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Summary")
    apply_sheet_title(sheet, "Sales Summary Logic")
    sheet.sheet_view.showGridLines = False

    labels = [
        ("A3", "Metric"),
        ("B3", "Value"),
        ("A4", "Total Revenue"),
        ("A5", "Parsed Orders"),
        ("A6", "Average Order Value"),
        ("A7", "Unique Customers"),
        ("A8", "Failed Parses"),
        ("A9", "Top Product"),
    ]
    for cell_ref, value in labels:
        sheet[cell_ref] = value
        sheet[cell_ref].font = Font(bold=True) if cell_ref.endswith("3") else Font(bold=False)

    sheet["B4"] = '=SUMIFS(RawData!$H$2:$H$5000,RawData!$J$2:$J$5000,"parsed")'
    sheet["B5"] = '=COUNTIF(RawData!$J$2:$J$5000,"parsed")'
    sheet["B6"] = '=IFERROR(B4/B5,0)'
    sheet["B7"] = "=SUM(RawData!$M$2:$M$5000)"
    sheet["B8"] = '=COUNTIF(RawData!$J$2:$J$5000,"failed")'
    sheet["B9"] = "=INDEX(F4:F9,MATCH(MAX(G4:G9),G4:G9,0))"

    for cell_ref in ["B4", "B6"]:
        sheet[cell_ref].number_format = '"$"#,##0;[Red]("$"#,##0)'

    metric_fill = PatternFill("solid", fgColor="EAF3F5")
    value_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_teal = Side(style="thin", color="B4CCD2")
    border = Border(bottom=thin_teal)

    for row in range(3, 10):
        for column in range(1, 3):
            cell = sheet.cell(row=row, column=column)
            cell.fill = metric_fill if column == 1 else value_fill
            cell.border = border

    sheet["A12"] = "Month"
    sheet["B12"] = "Revenue"
    sheet["C12"] = "Orders"
    for column in range(1, 4):
        sheet.cell(row=12, column=column).font = Font(bold=True)
        sheet.cell(row=12, column=column).fill = PatternFill("solid", fgColor="D9E7EB")

    sheet["A13"] = '=DATE(YEAR(TODAY()),MONTH(TODAY())-5,1)'
    for row in range(14, 19):
        sheet[f"A{row}"] = f"=DATE(YEAR(A{row-1}),MONTH(A{row-1})+1,1)"
    for row in range(13, 19):
        sheet[f"B{row}"] = (
            f'=SUMIFS(RawData!$H$2:$H$5000,RawData!$J$2:$J$5000,"parsed",'
            f'RawData!$K$2:$K$5000,$A{row})'
        )
        sheet[f"C{row}"] = (
            f'=COUNTIFS(RawData!$J$2:$J$5000,"parsed",RawData!$K$2:$K$5000,$A{row})'
        )
        sheet[f"A{row}"].number_format = "mmm-yy"
        sheet[f"B{row}"].number_format = '"$"#,##0;[Red]("$"#,##0)'

    sheet["F3"] = "Product"
    sheet["G3"] = "Revenue"
    sheet["H3"] = "Orders"
    for column in range(6, 9):
        sheet.cell(row=3, column=column).font = Font(bold=True)
        sheet.cell(row=3, column=column).fill = PatternFill("solid", fgColor="D9E7EB")

    product_list = [
        "MacBook Air",
        "iPhone 15",
        "Galaxy S24",
        "iPad Air",
        "Apple Watch",
        "AirPods Pro",
    ]
    for row, product in enumerate(product_list, start=4):
        sheet[f"F{row}"] = product
        sheet[f"G{row}"] = (
            f'=SUMIFS(RawData!$H$2:$H$5000,RawData!$J$2:$J$5000,"parsed",'
            f'RawData!$G$2:$G$5000,F{row})'
        )
        sheet[f"H{row}"] = (
            f'=COUNTIFS(RawData!$J$2:$J$5000,"parsed",RawData!$G$2:$G$5000,F{row})'
        )
        sheet[f"G{row}"].number_format = '"$"#,##0;[Red]("$"#,##0)'

    sheet["J3"] = "Status"
    sheet["K3"] = "Count"
    for column in range(10, 12):
        sheet.cell(row=3, column=column).font = Font(bold=True)
        sheet.cell(row=3, column=column).fill = PatternFill("solid", fgColor="D9E7EB")
    sheet["J4"] = "parsed"
    sheet["J5"] = "failed"
    sheet["K4"] = '=COUNTIF(RawData!$J$2:$J$5000,"parsed")'
    sheet["K5"] = '=COUNTIF(RawData!$J$2:$J$5000,"failed")'

    for column, width in {"A": 16, "B": 16, "C": 12, "F": 18, "G": 16, "H": 12, "J": 12, "K": 12}.items():
        sheet.column_dimensions[column].width = width


def style_kpi_card(
    sheet,
    start_cell: str,
    end_column: str,
    label: str,
    formula: str,
    number_format: str | None = None,
) -> None:
    label_cell = sheet[start_cell]
    start_column = get_column_letter(label_cell.column)
    value_row = label_cell.row + 1
    label_range = f"{start_column}{label_cell.row}:{end_column}{label_cell.row}"
    value_range = f"{start_column}{value_row}:{end_column}{value_row}"
    sheet.merge_cells(label_range)
    sheet.merge_cells(value_range)

    label_cell.value = label
    value_cell = sheet.cell(row=value_row, column=label_cell.column)
    value_cell.value = formula

    card_fill = PatternFill("solid", fgColor="EAF3F5")
    accent_fill = PatternFill("solid", fgColor="0B5D6B")
    border_side = Side(style="thin", color="9BB9C2")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    label_cell.fill = accent_fill
    label_cell.font = Font(color="FFFFFF", bold=True, size=11)
    label_cell.alignment = Alignment(horizontal="center", vertical="center")
    label_cell.border = border

    value_cell.fill = card_fill
    value_cell.font = Font(color="000000", bold=True, size=18)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")
    value_cell.border = border
    if number_format:
        value_cell.number_format = number_format

    for row in range(label_cell.row, value_row + 1):
        for column in range(label_cell.column, sheet[f"{end_column}{row}"].column + 1):
            sheet.cell(row=row, column=column).fill = accent_fill if row == label_cell.row else card_fill
            sheet.cell(row=row, column=column).border = border


def build_dashboard_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Dashboard")
    sheet.sheet_view.showGridLines = False
    apply_sheet_title(sheet, "WhatsApp Sales Dashboard")
    sheet.unmerge_cells("A1:H1")
    sheet.merge_cells("A1:L1")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.merge_cells("A2:L2")
    sheet["A2"] = "Built for the WhatsApp-to-Google-Sheets sales flow in this repo. Replace sample rows on RawData to refresh the view."
    sheet["A2"].font = Font(italic=True, color="5A5A5A")
    sheet["A2"].alignment = Alignment(wrap_text=True)

    style_kpi_card(
        sheet,
        "A4",
        "C",
        "Total Revenue",
        "=Summary!B4",
        '"$"#,##0;[Red]("$"#,##0)',
    )
    style_kpi_card(sheet, "D4", "F", "Parsed Orders", "=Summary!B5")
    style_kpi_card(
        sheet,
        "G4",
        "I",
        "Average Order Value",
        "=Summary!B6",
        '"$"#,##0;[Red]("$"#,##0)',
    )
    style_kpi_card(sheet, "J4", "L", "Unique Customers", "=Summary!B7")

    style_kpi_card(sheet, "A8", "C", "Failed Parses", "=Summary!B8")
    style_kpi_card(sheet, "D8", "F", "Top Product", "=Summary!B9")

    sheet.merge_cells("G8:L10")
    sheet["G8"] = (
        "Sources:\n"
        "https://web.whatsapp.com/\n"
        "https://developers.google.com/sheets/api\n"
        "Schema aligned to the local project README."
    )
    sheet["G8"].fill = PatternFill("solid", fgColor="F5F7F8")
    sheet["G8"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet["G8"].font = Font(size=10, color="4A4A4A")
    note_border_side = Side(style="thin", color="D0D9DD")
    note_border = Border(
        left=note_border_side,
        right=note_border_side,
        top=note_border_side,
        bottom=note_border_side,
    )
    for row in range(8, 11):
        for column in range(7, 13):
            sheet.cell(row=row, column=column).fill = PatternFill("solid", fgColor="F5F7F8")
            sheet.cell(row=row, column=column).border = note_border

    monthly_chart = LineChart()
    monthly_chart.title = "Revenue Trend (Last 6 Months)"
    monthly_chart.style = 2
    monthly_chart.y_axis.title = "Revenue"
    monthly_chart.x_axis.title = "Month"
    monthly_data = Reference(workbook["Summary"], min_col=2, min_row=12, max_row=18)
    monthly_cats = Reference(workbook["Summary"], min_col=1, min_row=13, max_row=18)
    monthly_chart.add_data(monthly_data, titles_from_data=True)
    monthly_chart.set_categories(monthly_cats)
    monthly_chart.height = 7.2
    monthly_chart.width = 11.2
    sheet.add_chart(monthly_chart, "A13")

    product_chart = BarChart()
    product_chart.type = "bar"
    product_chart.style = 10
    product_chart.title = "Revenue by Product"
    product_chart.x_axis.title = "Revenue"
    product_chart.y_axis.title = "Product"
    product_data = Reference(workbook["Summary"], min_col=7, min_row=3, max_row=9)
    product_cats = Reference(workbook["Summary"], min_col=6, min_row=4, max_row=9)
    product_chart.add_data(product_data, titles_from_data=True)
    product_chart.set_categories(product_cats)
    product_chart.height = 7.0
    product_chart.width = 9.6
    sheet.add_chart(product_chart, "G13")

    status_chart = PieChart()
    status_chart.title = "Parse Status Mix"
    status_data = Reference(workbook["Summary"], min_col=11, min_row=3, max_row=5)
    status_labels = Reference(workbook["Summary"], min_col=10, min_row=4, max_row=5)
    status_chart.add_data(status_data, titles_from_data=True)
    status_chart.set_categories(status_labels)
    status_chart.height = 6.2
    status_chart.width = 6.8
    sheet.add_chart(status_chart, "J29")

    column_widths = {
        "A": 14,
        "B": 14,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 14,
        "H": 14,
        "I": 14,
        "J": 14,
        "K": 14,
        "L": 14,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width
    sheet.row_dimensions[2].height = 34
    for row in [4, 8]:
        sheet.row_dimensions[row].height = 22
        sheet.row_dimensions[row + 1].height = 34
    for row in range(13, 38):
        sheet.row_dimensions[row].height = 20


def build_instructions_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Instructions")
    apply_sheet_title(sheet, "How To Use This Dashboard")
    sheet.sheet_view.showGridLines = False

    instructions = [
        "1. Replace the example rows on RawData with exports from your Google Sheet.",
        "2. Keep the column order from A:J unchanged so the formulas continue to work.",
        "3. Columns K:M are helper formulas; leave them in place and copy them downward for new rows if Excel does not autofill them.",
        "4. If you introduce new product names, update Summary!F4:F9 so the product chart reflects your catalog.",
        "5. Open the workbook in Excel or Google Sheets to refresh formulas and chart values.",
    ]
    for row, text in enumerate(instructions, start=3):
        sheet[f"A{row}"] = text
        sheet[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 110


def create_workbook() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    rows = sample_sales_data()
    build_raw_data_sheet(workbook, rows)
    build_summary_sheet(workbook)
    build_dashboard_sheet(workbook)
    build_instructions_sheet(workbook)

    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    path = create_workbook()
    print(path)
